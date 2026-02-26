"""
AUDITORIA DE CERTIFICADOS SSL v5.0
- Sin colores de comparacion en Excel (solo en log)
- Amarillo = vence en menos de DIAS_ALERTA dias
- Rojo = ya vencido
- Hojas WAS: fecha en col F (signer) o col G (personal)
- Hojas AIPAC: fecha en col K
- Hojas PLUG.WAS: sin fechas de certificado
- Argparse para configuracion por CLI
- Cache de directorio para mejora de rendimiento
- Log con fecha completa (YYYY-MM-DD HH:MM:SS)
- Estadisticas de cobertura al finalizar
- HTML con exportacion CSV y paginacion
"""

import os
import re
import glob
import argparse
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ==========================
# ARGUMENTOS CLI
# ==========================
def parse_args():
    parser = argparse.ArgumentParser(
        description="Auditoria de Certificados SSL v5.0 - COMBMAN Keystores"
    )
    parser.add_argument(
        "--raiz",
        default=r"C:\Automatizacion_Excel",
        help="Carpeta raiz del proyecto (default: C:\\Automatizacion_Excel)"
    )
    parser.add_argument(
        "--dias-alerta",
        type=int,
        default=90,
        help="Dias antes del vencimiento para alerta amarilla (default: 90)"
    )
    parser.add_argument(
        "--excel-in",
        default=None,
        help="Ruta al Excel de entrada (default: RAIZ/REPORTE_AUDITORIA.xlsx)"
    )
    parser.add_argument(
        "--ambientes",
        nargs="+",
        default=["CAMARAPROD", "CAMARARESP", "CAMARATEST"],
        help="Lista de ambientes a procesar"
    )
    return parser.parse_args()


# ==========================
# CONFIGURACION (poblada desde args)
# ==========================
_args = parse_args()

RAIZ         = _args.raiz
CARPETA_BASE = os.path.join(RAIZ, "PROCESADOS")
EXCEL_IN     = _args.excel_in or os.path.join(RAIZ, "REPORTE_AUDITORIA.xlsx")
ARCHIVO_LOG      = os.path.join(RAIZ, "LOG_PROCESAMIENTO.txt")
LOG_VENCIMIENTOS = os.path.join(RAIZ, "LOG_VENCIMIENTOS.txt")
HTML_REPORTE     = os.path.join(RAIZ, "REPORTE_AUDITORIA.html")

AMBIENTES   = _args.ambientes
DIAS_ALERTA = _args.dias_alerta

# Nombre del archivo de salida con mes anterior al de ejecucion
def _nombre_excel_salida():
    hoy = date.today()
    if hoy.month == 1:
        mes_ant = 12
        anio    = hoy.year - 1
    else:
        mes_ant = hoy.month - 1
        anio    = hoy.year
    MESES_NOMBRE = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    return os.path.join(RAIZ, f"COMBMAN. Keystores de Infraestructura y Seguridad - {MESES_NOMBRE[mes_ant]} {anio}.xlsx")


FILL_VENCIDO = PatternFill("solid", fgColor="FF0000")   # rojo  = vencido
FILL_PROXIMO = PatternFill("solid", fgColor="FFFF00")   # amarillo = por vencer
FILL_OK      = PatternFill(fill_type=None)              # sin color = vigente


# ==========================
# LOG
# ==========================
def log(msg, nivel="INFO"):
    ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linea = f"[{ts}] [{nivel}] {msg}"
    print(linea)
    with open(ARCHIVO_LOG, "a", encoding="utf-8") as f:
        f.write(linea + "\n")


# ==========================
# ESTADISTICAS DE COBERTURA
# ==========================
_stats = {"resueltos": 0, "no_encontrados": 0, "total_aliases": 0}

def stats_resuelto():
    _stats["resueltos"]     += 1
    _stats["total_aliases"] += 1

def stats_no_encontrado():
    _stats["no_encontrados"] += 1
    _stats["total_aliases"]  += 1

def imprimir_estadisticas():
    total = _stats["total_aliases"]
    res   = _stats["resueltos"]
    nf    = _stats["no_encontrados"]
    pct   = (res / total * 100) if total > 0 else 0
    log("=" * 60)
    log(f"  ESTADISTICAS DE COBERTURA")
    log(f"  Total aliases procesados : {total}")
    log(f"  Resueltos (con .out)     : {res}  ({pct:.1f}%)")
    log(f"  Sin archivo .out         : {nf}")
    log("=" * 60)


# ==========================
# PARSEO DE FECHAS
# ==========================
MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}
MESES_EN = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "mary": 5   # typo comun en el excel
}
MESES = {**MESES_ES, **MESES_EN}


def extraer_fecha_vencimiento(texto):
    """
    Extrae la fecha de vencimiento de strings como:
      'Valid from May 18, 2025 to May 18, 2026.'
      'Valid from 18 Mayo 2025, to  18 Mayo 2026'
      '10/26/07 7:42 AM until: 10/21/27 7:42 AM'
      'Valid from may 2025, to May 14 2040'
      '2026-05-18'  (ISO)
    Retorna un objeto date o None.
    """
    if not texto:
        return None
    texto = str(texto).strip()

    # Formato ISO: YYYY-MM-DD (agregado en v5.0)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", texto)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass

    # Formato: MM/DD/YY ... until: MM/DD/YY
    m = re.search(r"until:\s*(\d{1,2})/(\d{1,2})/(\d{2,4})", texto, re.IGNORECASE)
    if m:
        mes, dia, anio = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if anio < 100:
            anio += 2000
        try:
            return date(anio, mes, dia)
        except Exception:
            pass

    # Formato: "to Month DD, YYYY" o "to DD Month YYYY"
    patrones_fin = [
        r"to\s+(\w+)\s+(\d{1,2}),?\s+(\d{4})",    # to May 18, 2026
        r"to\s+(\d{1,2})\s+(\w+)\s+(\d{4})",       # to 18 Mayo 2026
        r"to\s+(\w+)\s+(\d{1,2})\s+(\d{4})",       # to May 18 2026
        r"to\s+(\w+)\s+(\d{4})",                    # to May 2040 (sin dia)
    ]

    for patron in patrones_fin:
        m = re.search(patron, texto, re.IGNORECASE)
        if m:
            g1, g2, g3 = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()

            # Intentar "to Mes Dia Anio"
            if g1.lower() in MESES:
                mes = MESES[g1.lower()]
                try:
                    dia  = int(g2)
                    anio = int(g3)
                    return date(anio, mes, dia)
                except Exception:
                    pass

            # Intentar "to Dia Mes Anio"
            if g2.lower() in MESES:
                try:
                    dia  = int(g1)
                    mes  = MESES[g2.lower()]
                    anio = int(g3)
                    return date(anio, mes, dia)
                except Exception:
                    pass

    # Formato: "December 31, 2028" al final
    m = re.search(r"(\w+)\s+(\d{1,2}),?\s+(\d{4})\s*[.\s]*$", texto, re.IGNORECASE)
    if m:
        mes_str = m.group(1).lower()
        if mes_str in MESES:
            try:
                return date(int(m.group(3)), MESES[mes_str], int(m.group(2)))
            except Exception:
                pass

    return None


def evaluar_vencimiento(fecha_venc, alias, hoja):
    """Evalua estado del certificado y retorna (fill, mensaje_log)."""
    if fecha_venc is None:
        return None, f"    [{hoja}] '{alias}': fecha no parseable"

    hoy       = date.today()
    dias_rest = (fecha_venc - hoy).days

    if dias_rest < 0:
        return FILL_VENCIDO, f"    [{hoja}] '{alias}': VENCIDO hace {abs(dias_rest)} dias ({fecha_venc})"
    elif dias_rest <= DIAS_ALERTA:
        return FILL_PROXIMO, f"    [{hoja}] '{alias}': PROXIMO A VENCER en {dias_rest} dias ({fecha_venc})"
    else:
        return FILL_OK, f"    [{hoja}] '{alias}': vigente ({dias_rest} dias restantes, {fecha_venc})"


# ==========================
# PARSERS ARCHIVOS .out
# ==========================
def parsear_out(ruta):
    try:
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            texto = f.read()
        label  = re.search(r"^Label\s*:\s*(.+)$", texto, re.MULTILINE)
        serial = re.search(r"^Serial\s*:\s*(.+)$", texto, re.MULTILINE)
        sha1   = re.search(r"Fingerprint\s*:\s*SHA1\s*:\s*\n([\s\S]*?)(?=Fingerprint\s*:|$)", texto)
        if not label or not serial or not sha1:
            return None
        bytes_hex = re.findall(r"[0-9A-Fa-f]{2}", sha1.group(1))
        return {
            "label":  label.group(1).strip().lower(),
            "serial": serial.group(1).strip().lower(),
            "sha1":   " ".join(bytes_hex).upper(),
        }
    except Exception as e:
        log(f"Error parseando {ruta}: {e}", "ERROR")
        return None


def parsear_out_keytool(ruta):
    resultado = {}
    try:
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            texto = f.read()
        bloques = re.split(r"\n(?=Alias name:)", texto)
        for bloque in bloques:
            alias_m  = re.search(r"^Alias name:\s*(.+)$", bloque, re.MULTILINE)
            serial_m = re.search(r"^Serial number:\s*([0-9a-fA-F]+)", bloque, re.MULTILINE)
            sha1_m   = re.search(r"SHA1:\s*([0-9A-Fa-f:]+)", bloque)
            if alias_m and serial_m:
                alias = alias_m.group(1).strip().lower()
                resultado[alias] = {
                    "serial": serial_m.group(1).strip().lower(),
                    "sha1":   sha1_m.group(1).strip() if sha1_m else "",
                }
    except Exception as e:
        log(f"Error parseando keytool {ruta}: {e}", "ERROR")
    return resultado


def parsear_sha256(ruta):
    mapa = {}
    try:
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            for linea in f:
                linea = linea.strip()
                if not linea:
                    continue
                partes = linea.split(None, 1)
                if len(partes) == 2:
                    mapa[partes[1].strip()] = partes[0].strip()
    except Exception as e:
        log(f"Error parseando sha256 {ruta}: {e}", "ERROR")
    return mapa


# ==========================
# BUSQUEDA DE ARCHIVOS (con cache)
# ==========================
_dir_cache = {}

def _listar_carpeta(carpeta):
    """Lista el contenido de una carpeta con cache para evitar listdir repetido."""
    if carpeta not in _dir_cache:
        if os.path.exists(carpeta):
            _dir_cache[carpeta] = os.listdir(carpeta)
        else:
            _dir_cache[carpeta] = []
    return _dir_cache[carpeta]


def carpeta_ambiente(ambiente):
    base = CARPETA_BASE
    if os.path.exists(base):
        for d in os.listdir(base):
            if d.upper() == ambiente.upper():
                return os.path.join(base, d)
    return os.path.join(base, ambiente)


def alias_a_nombre(alias):
    nombre = str(alias).strip()
    nombre = re.sub(r"\s+", "-", nombre)
    return nombre.strip("-")


def _similitud_alias(alias, nombre_archivo):
    """
    Compara alias del Excel con nombre del archivo.
    Normaliza ambos quitando espacios, guiones, puntos, prefijos SC_/PC_
    y compara si uno contiene al otro o son suficientemente similares.
    """
    def normalizar(s):
        s = re.sub(r"^(SC_|PC_)", "", s, flags=re.IGNORECASE)
        s = re.sub(r"[\s\-_=,\.]+", "", s).lower()
        return s

    a = normalizar(alias)
    # Extraer solo la parte del alias del nombre de archivo (quitar ambiente_num_SC_)
    partes = nombre_archivo.replace(".out", "").split("_", 3)
    b = normalizar(partes[-1]) if len(partes) >= 4 else normalizar(nombre_archivo)

    # Coincidencia exacta normalizada
    if a == b:
        return True
    # Uno contiene al otro (alias corto vs nombre largo)
    if len(a) >= 4 and (a in b or b in a):
        return True
    # Sin la ultima letra (typos)
    if len(a) >= 4 and (a[:-1] == b or a == b[:-1]):
        return True
    return False


def buscar_out_alias(ambiente, numero, alias, carpeta):
    alias_norm   = str(alias).strip().lower()
    alias_limpio = re.sub(r"^(SC_|PC_)", "", str(alias).strip(), flags=re.IGNORECASE)
    alias_archivo = alias_a_nombre(alias_limpio)

    # 1. Intento exacto con SC y PC
    variaciones = [alias_archivo]
    if len(alias_archivo) > 3:
        variaciones.append(alias_archivo[:-1])

    for tipo in ["SC", "PC"]:
        for var in variaciones:
            ruta = os.path.join(carpeta, f"{ambiente.lower()}_{numero}_{tipo}_{var}.out")
            if os.path.exists(ruta):
                datos = parsear_out(ruta)
                if datos:
                    return ruta, datos

    # 2. Busqueda flexible usando cache de directorio
    prefijo    = f"{ambiente.lower()}_{numero}_"
    candidatos = [f for f in _listar_carpeta(carpeta)
                  if f.lower().startswith(prefijo) and f.lower().endswith(".out")]

    # Primero: comparar Label interno
    for fname in candidatos:
        ruta  = os.path.join(carpeta, fname)
        datos = parsear_out(ruta)
        if datos and datos["label"] == alias_norm:
            return ruta, datos

    # Segundo: comparar por similitud de nombre de archivo
    for fname in candidatos:
        if _similitud_alias(alias, fname):
            ruta  = os.path.join(carpeta, fname)
            datos = parsear_out(ruta)
            if datos:
                return ruta, datos

    return None, None


def buscar_keystore_out(ambiente, nombre_ks, carpeta):
    for f in _listar_carpeta(carpeta):
        if f.lower().startswith(ambiente.lower()) and nombre_ks.lower() in f.lower():
            return os.path.join(carpeta, f)
    return None


def buscar_sha256(ambiente, carpeta):
    for f in _listar_carpeta(carpeta):
        if f.lower().startswith(ambiente.lower()) and "plugwas" in f.lower():
            return os.path.join(carpeta, f)
    return None


# ==========================
# NORMALIZACION
# ==========================
def norm_fp(valor):
    if not valor:
        return ""
    v = re.sub(r"^SHA\d+\s*:\s*", "", str(valor), flags=re.IGNORECASE)
    return " ".join(re.findall(r"[0-9A-Fa-f]{2}", v)).upper()


def norm_serial(valor):
    if not valor:
        return ""
    return re.sub(r"[\s\-:]", "", str(valor)).lower()


ALIAS_SKIP = {
    "alias", "no existen", "# signer certificates",
    "# personal certificates", "# personal certificate requests",
    "# custom properties", "issued by", "issued to",
    "keystore provider: ibmjce"
}

def es_alias_valido(alias):
    if not alias:
        return False
    a = str(alias).strip().lower()
    return a not in ALIAS_SKIP


# ==========================
# PROCESAR HOJA WAS
# ==========================
def procesar_hoja_was(ws, ambiente, diffs):
    carpeta = carpeta_ambiente(ambiente)
    log(f"  -> Procesando hoja WAS: {ws.title}")
    seccion_actual = None
    modo_personal  = False  # False=signer(col F), True=personal(col G)

    for row in ws.iter_rows(min_row=1):
        col_a = row[0].value
        col_b = row[1].value
        col_c = row[2].value
        col_f = row[5]   # Expiration signer certs
        col_g = row[6]   # Expiration personal certs

        # Detectar seccion
        if col_a and re.match(r"^\d+\.-", str(col_a).strip()):
            m = re.match(r"^(\d+)\.-", str(col_a).strip())
            seccion_actual = int(m.group(1))
            modo_personal  = False
            log(f"    Seccion {seccion_actual}: {col_b}")
            continue

        # Detectar subseccion personal certificates
        if col_b and "# personal certificates" in str(col_b).lower():
            modo_personal = True
            continue

        if not es_alias_valido(col_c):
            continue

        alias = str(col_c).strip()

        # Celda de fecha segun modo
        fecha_cell = col_g if modo_personal else col_f

        # Colorear segun vencimiento
        fecha_venc = extraer_fecha_vencimiento(fecha_cell.value)
        fill, msg  = evaluar_vencimiento(fecha_venc, alias, ws.title)
        log(msg, "VENC" if fill == FILL_VENCIDO else ("ALERT" if fill == FILL_PROXIMO else "INFO"))
        if fill is not None:
            fecha_cell.fill = fill
        if fill in (FILL_VENCIDO, FILL_PROXIMO):
            clave = f"{ws.title}|{alias}|{fill}"
            if clave not in _diffs_set:
                _diffs_set.add(clave)
                diffs.append(msg.strip())

        # Comparar datos con .out
        if not seccion_actual:
            continue

        fp_val     = row[4].value  # col E
        serial_val = row[5].value  # col F (cuando es personal, serial en F)

        tiene_fp     = fp_val     and bool(re.search(r"[0-9A-Fa-f]{2}[\s:]", str(fp_val)))
        tiene_serial = not modo_personal and serial_val and bool(re.match(r"^[0-9a-fA-F\s]{4,}$", str(serial_val).strip()))
        if modo_personal:
            serial_val   = row[5].value
            tiene_serial = serial_val and bool(re.match(r"^[0-9a-fA-F\s]{4,}$", str(serial_val).strip()))
            tiene_fp     = False

        _, datos = buscar_out_alias(ambiente, seccion_actual, alias, carpeta)

        if not datos:
            log(f"    [{ws.title}] #{seccion_actual} '{alias}': .out no encontrado", "WARN")
            stats_no_encontrado()
            continue

        stats_resuelto()

        if tiene_fp:
            fp_excel = norm_fp(fp_val)
            fp_out   = norm_fp(datos["sha1"])
            if fp_excel == fp_out:
                log(f"    #{seccion_actual} '{alias}' FP: IGUAL")
            else:
                log(f"    #{seccion_actual} '{alias}' FP: DIFERENTE -> actualizando", "CAMBIO")
                diffs.append(f"{ws.title} | #{seccion_actual} {alias} | FP actualizado")
                row[4].value = datos["sha1"]

        if tiene_serial:
            s_excel = norm_serial(serial_val)
            s_out   = norm_serial(datos["serial"])
            if s_excel == s_out:
                log(f"    #{seccion_actual} '{alias}' Serial: IGUAL")
            else:
                log(f"    #{seccion_actual} '{alias}' Serial: DIFERENTE -> actualizando", "CAMBIO")
                diffs.append(f"{ws.title} | #{seccion_actual} {alias} | Serial actualizado")
                row[5].value = datos["serial"]


# ==========================
# PROCESAR HOJA PLUG.WAS
# ==========================
def procesar_hoja_plug_was(ws, ambiente, diffs):
    carpeta = carpeta_ambiente(ambiente)
    log(f"  -> Procesando hoja PLUG.WAS: {ws.title}")

    ruta_sha256 = buscar_sha256(ambiente, carpeta)
    if not ruta_sha256:
        log(f"    [{ws.title}] Archivo .sha256 no encontrado para {ambiente}", "WARN")
        return

    mapa = parsear_sha256(ruta_sha256)
    log(f"    sha256 cargado: {os.path.basename(ruta_sha256)} ({len(mapa)} rutas)")

    for row in ws.iter_rows(min_row=2):
        cell_ruta = row[0]
        cell_hash = row[1]
        ruta_excel = str(cell_ruta.value).strip() if cell_ruta.value else ""
        hash_excel = str(cell_hash.value).strip() if cell_hash.value else ""

        if not ruta_excel or not ruta_excel.startswith("/"):
            continue

        hash_out = mapa.get(ruta_excel)
        if hash_out is None:
            log(f"    '{ruta_excel}': no en .sha256", "WARN")
            continue

        if hash_excel.lower() == hash_out.lower():
            log(f"    '{ruta_excel}': IGUAL")
        else:
            log(f"    '{ruta_excel}': DIFERENTE -> actualizando", "CAMBIO")
            diffs.append(f"{ws.title} | {ruta_excel} | hash actualizado")
            cell_hash.value = hash_out


# ==========================
# PROCESAR HOJA AIPAC
# ==========================
def procesar_hoja_aipac(ws, ambiente, diffs):
    carpeta = carpeta_ambiente(ambiente)
    log(f"  -> Procesando hoja AIPAC: {ws.title}")

    mapa_ks = {}
    for ks in ["DSkeystore", "SSLkeystore"]:
        ruta = buscar_keystore_out(ambiente, ks, carpeta)
        if ruta:
            datos = parsear_out_keytool(ruta)
            mapa_ks[ks.lower()] = datos
            log(f"    {ks} cargado: {len(datos)} aliases")
        else:
            log(f"    {ks}: no encontrado para {ambiente}", "WARN")
            mapa_ks[ks.lower()] = {}

    seccion_ks = None
    for row in ws.iter_rows(min_row=1):
        col_b  = str(row[1].value).strip() if row[1].value else ""
        col_c  = row[2]
        col_j  = row[9]   # Serial number  (col J)
        col_k  = row[10]  # Expiration     (col K)

        if "dskeystore" in col_b.lower():
            seccion_ks = "dskeystore"
            log(f"    Seccion DSkeystore")
            continue
        if "sslkeystore" in col_b.lower():
            seccion_ks = "sslkeystore"
            log(f"    Seccion SSLkeystore")
            continue

        if not es_alias_valido(col_c.value):
            continue

        alias = str(col_c.value).strip()

        # Colorear segun vencimiento (col K)
        fecha_venc = extraer_fecha_vencimiento(col_k.value)
        fill, msg  = evaluar_vencimiento(fecha_venc, alias, ws.title)
        log(msg, "VENC" if fill == FILL_VENCIDO else ("ALERT" if fill == FILL_PROXIMO else "INFO"))
        if fill is not None:
            col_k.fill = fill
        if fill in (FILL_VENCIDO, FILL_PROXIMO):
            clave = f"{ws.title}|{alias}|{fill}"
            if clave not in _diffs_set:
                _diffs_set.add(clave)
                diffs.append(msg.strip())

        # Comparar Serial
        serial_val = col_j.value
        if not serial_val or not seccion_ks:
            continue

        datos = mapa_ks.get(seccion_ks, {}).get(alias.lower())
        if not datos:
            log(f"    [{seccion_ks}] '{alias}': no en .out", "WARN")
            stats_no_encontrado()
            continue

        stats_resuelto()
        s_excel = norm_serial(serial_val)
        s_out   = norm_serial(datos["serial"])
        if s_excel == s_out:
            log(f"    [{seccion_ks}] '{alias}' Serial: IGUAL")
        else:
            log(f"    [{seccion_ks}] '{alias}' Serial: DIFERENTE -> actualizando", "CAMBIO")
            diffs.append(f"{ws.title} | {seccion_ks} | {alias} | Serial actualizado")
            col_j.value = datos["serial"]


# ==========================
# GENERADOR HTML
# ==========================
def generar_html_reporte(archivo_log, archivo_html, fecha_ejecucion, dias_alerta):
    """Lee el LOG_PROCESAMIENTO.txt y genera un reporte HTML interactivo."""
    import json

    registros = []

    try:
        with open(archivo_log, "r", encoding="utf-8", errors="ignore") as f:
            lineas = f.readlines()
    except Exception:
        return

    hoja_actual = "-"
    _vistos_act = set()
    for linea in lineas:
        linea = linea.strip()

        # Rastrear hoja activa del log
        mh = re.search(r'Hoja: (\S+)', linea)
        if mh:
            hoja_actual = mh.group(1)

        # VENCIDO
        m = re.search(r"\[VENC\].*?\[(.+?)\] '(.+?)': VENCIDO hace (\d+) dias \((.+?)\)", linea)
        if m:
            registros.append({"hoja": m.group(1), "alias": m.group(2),
                               "estado": "VENCIDO", "dias": -int(m.group(3)), "fecha": m.group(4), "detalle": ""})
            continue

        # PROXIMO A VENCER
        m = re.search(r"\[ALERT\].*?\[(.+?)\] '(.+?)': PROXIMO A VENCER en (\d+) dias \((.+?)\)", linea)
        if m:
            existe = any(r["hoja"] == m.group(1) and r["alias"] == m.group(2)
                        and r["estado"] == "PROXIMO" for r in registros)
            if not existe:
                registros.append({"hoja": m.group(1), "alias": m.group(2),
                                   "estado": "PROXIMO", "dias": int(m.group(3)), "fecha": m.group(4), "detalle": ""})
            continue

        # CAMBIO WAS/AIPAC - resumen final: "HOJA | #N alias | FP/Serial actualizado"
        m = re.search(r"\[CAMBIO\]\s+(\S+-(?:WAS|AIPAC))\s+\|\s+#?\w*\s*(.+?)\s+\|\s+(FP|Serial) actualizado", linea)
        if m:
            key = (m.group(1), m.group(2).strip(), "ACTUALIZADO")
            if key not in _vistos_act:
                _vistos_act.add(key)
                registros.append({"hoja": m.group(1), "alias": m.group(2).strip(), "estado": "ACTUALIZADO",
                                   "dias": 9999, "fecha": "-", "detalle": m.group(3) + " actualizado"})
            continue

        # CAMBIO AIPAC keystore - resumen final: "HOJA | keystore | alias | Serial actualizado"
        m = re.search(r"\[CAMBIO\]\s+(\S+-AIPAC)\s+\|\s+(\w+)\s+\|\s+(.+?)\s+\|\s+(Serial) actualizado", linea)
        if m:
            key = (m.group(1), m.group(3).strip(), "ACTUALIZADO")
            if key not in _vistos_act:
                _vistos_act.add(key)
                registros.append({"hoja": m.group(1), "alias": m.group(3).strip() + " (" + m.group(2) + ")",
                                   "estado": "ACTUALIZADO", "dias": 9999, "fecha": "-",
                                   "detalle": m.group(4) + " actualizado"})
            continue

        # CAMBIO PLUG.WAS - resumen final: "HOJA-COMP.PLUG.WAS | ruta | hash actualizado"
        m = re.search(r"\[CAMBIO\]\s+(\S+-COMP\.PLUG\.WAS)\s+\|\s+(.+?)\s+\|\s+hash actualizado", linea)
        if m:
            # Mostrar solo el nombre del archivo, no la ruta completa
            ruta = m.group(2).strip()
            alias = ruta.split("/")[-1] if "/" in ruta else ruta
            key = (m.group(1), alias, "ACTUALIZADO")
            if key not in _vistos_act:
                _vistos_act.add(key)
                registros.append({"hoja": m.group(1), "alias": alias, "estado": "ACTUALIZADO",
                                   "dias": 9999, "fecha": "-", "detalle": "SHA256 actualizado"})
            continue

        # CAMBIO WAS durante procesamiento (fallback con hoja_actual)
        m = re.search(r"\[CAMBIO\].*?#?(\d+)? ?'(.+?)' (FP|Serial): DIFERENTE", linea)
        if m:
            key = (hoja_actual, m.group(2), "ACTUALIZADO")
            if key not in _vistos_act:
                _vistos_act.add(key)
                registros.append({"hoja": hoja_actual, "alias": m.group(2), "estado": "ACTUALIZADO",
                                   "dias": 9999, "fecha": "-", "detalle": m.group(3) + " actualizado"})
            continue

        # SIN ARCHIVO
        m = re.search(r"\[WARN\].*?\[(.+?)\] #(\d+) '(.+?)': \.out no encontrado", linea)
        if m:
            registros.append({"hoja": m.group(1), "alias": m.group(3),
                               "estado": "SIN_ARCHIVO", "dias": 9999, "fecha": "-",
                               "detalle": "Secc. " + m.group(2)})
            continue

    # Derivar ambiente de hoja
    for r in registros:
        for amb in AMBIENTES:
            if amb in r["hoja"]:
                r["ambiente"] = amb
                break
        else:
            r["ambiente"] = "-"

    # Contar por estado
    cnt = {"VENCIDO": 0, "PROXIMO": 0, "ACTUALIZADO": 0, "SIN_ARCHIVO": 0}
    for r in registros:
        if r["estado"] in cnt:
            cnt[r["estado"]] += 1

    total_con_estado = cnt["VENCIDO"] + cnt["PROXIMO"]
    total_certs      = max(total_con_estado + cnt["SIN_ARCHIVO"], 1)
    pct_critico      = round((cnt["VENCIDO"] + cnt["PROXIMO"]) / total_certs * 100)

    datos_json = json.dumps([{
        "ambiente": r.get("ambiente", "-"),
        "hoja":     r.get("hoja", "-"),
        "alias":    r.get("alias", "-"),
        "estado":   r.get("estado", "-"),
        "dias":     r.get("dias", 9999),
        "fecha":    r.get("fecha", "-"),
        "detalle":  r.get("detalle", ""),
    } for r in registros], ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Auditoria SSL - {fecha_ejecucion}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f0f2f5; color: #222; padding: 24px; }}
h1 {{ font-size: 20px; font-weight: 700; color: #1a1a2e; margin-bottom: 4px; }}
.subtitle {{ font-size: 13px; color: #666; margin-bottom: 20px; }}
/* Barra de criticidad */
.criticidad-wrap {{ margin-bottom: 24px; background: white; border-radius: 10px;
                    padding: 14px 18px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
.criticidad-label {{ font-size: 12px; font-weight: 600; color: #555; margin-bottom: 8px; }}
.criticidad-bar-bg {{ background: #e2e8f0; border-radius: 20px; height: 14px; overflow: hidden; }}
.criticidad-bar {{ height: 14px; border-radius: 20px; transition: width .4s;
                   background: linear-gradient(90deg, #f6ad55, #e53e3e); }}
.criticidad-pct {{ font-size: 13px; font-weight: 700; margin-top: 6px;
                   color: {'#e53e3e' if pct_critico > 30 else '#d69e2e' if pct_critico > 10 else '#38a169'}; }}
/* Cards */
.cards {{ display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }}
.card {{ background: white; border-radius: 10px; padding: 16px 22px; flex: 1; min-width: 160px;
         box-shadow: 0 1px 4px rgba(0,0,0,0.08); border-left: 4px solid #ccc; cursor: pointer; transition: transform .1s; }}
.card:hover {{ transform: translateY(-2px); }}
.card.rojo  {{ border-color: #e53e3e; }}
.card.amari {{ border-color: #d69e2e; }}
.card.azul  {{ border-color: #3182ce; }}
.card.gris  {{ border-color: #718096; }}
.card .num  {{ font-size: 32px; font-weight: 800; line-height: 1; margin-bottom: 4px; }}
.card.rojo  .num {{ color: #e53e3e; }}
.card.amari .num {{ color: #d69e2e; }}
.card.azul  .num {{ color: #3182ce; }}
.card.gris  .num {{ color: #718096; }}
.card .lbl  {{ font-size: 12px; color: #666; font-weight: 500; }}
/* Filtros */
.filtros {{ display: flex; gap: 10px; margin-bottom: 16px; flex-wrap: wrap; align-items: center; }}
.filtros label {{ font-size: 13px; font-weight: 600; color: #444; }}
select, input {{ padding: 7px 12px; border: 1px solid #d1d5db; border-radius: 6px;
                 font-size: 13px; background: white; cursor: pointer; }}
input {{ width: 220px; }}
.btn {{ padding: 7px 14px; border: 1px solid #d1d5db; border-radius: 6px;
        font-size: 13px; cursor: pointer; }}
.btn-reset {{ background: #f7fafc; color: #555; }}
.btn-reset:hover {{ background: #edf2f7; }}
.btn-csv {{ background: #2b6cb0; color: white; border-color: #2b6cb0; font-weight: 600; }}
.btn-csv:hover {{ background: #2c5282; }}
/* Tabla */
.tabla-wrap {{ background: white; border-radius: 10px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); overflow: hidden; }}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
thead {{ background: #1a1a2e; color: white; }}
th {{ padding: 12px 14px; text-align: left; font-weight: 600; font-size: 12px;
      text-transform: uppercase; letter-spacing: 0.5px; cursor: pointer; user-select: none; white-space: nowrap; }}
th:hover {{ background: #2d2d4e; }}
td {{ padding: 10px 14px; border-bottom: 1px solid #f0f2f5; vertical-align: middle; }}
tr:last-child td {{ border-bottom: none; }}
tr:hover td {{ background: #f8fafc; }}
.badge {{ display: inline-block; padding: 3px 10px; border-radius: 20px;
          font-size: 11px; font-weight: 700; letter-spacing: 0.3px; white-space: nowrap; }}
.badge-VENCIDO     {{ background: #fed7d7; color: #9b2335; }}
.badge-PROXIMO     {{ background: #fefcbf; color: #744210; }}
.badge-ACTUALIZADO {{ background: #bee3f8; color: #2b6cb0; }}
.badge-SIN_ARCHIVO {{ background: #e2e8f0; color: #4a5568; }}
tr.row-VENCIDO td  {{ background: #fff5f5; }}
tr.row-PROXIMO td  {{ background: #fffff0; }}
.dias-critico {{ color: #e53e3e; font-weight: 700; }}
.dias-alerta  {{ color: #d69e2e; font-weight: 700; }}
.dias-ok      {{ color: #38a169; }}
.no-rows {{ text-align: center; padding: 32px; color: #999; font-size: 14px; }}
/* Paginacion */
.paginacion {{ display: flex; gap: 6px; justify-content: center; align-items: center;
               margin-top: 14px; flex-wrap: wrap; }}
.paginacion button {{ padding: 5px 12px; border: 1px solid #d1d5db; border-radius: 6px;
                      font-size: 13px; background: white; cursor: pointer; }}
.paginacion button.activa {{ background: #1a1a2e; color: white; border-color: #1a1a2e; font-weight: 700; }}
.paginacion button:hover:not(.activa) {{ background: #edf2f7; }}
.paginacion .info {{ font-size: 12px; color: #888; margin: 0 8px; }}
.footer {{ margin-top: 12px; font-size: 12px; color: #999; text-align: right; }}
</style>
</head>
<body>
<h1>🔐 Auditoria SSL — COMBMAN Keystores de Infraestructura y Seguridad</h1>
<div class="subtitle">Generado: {fecha_ejecucion} &nbsp;|&nbsp; Umbral de alerta: {dias_alerta} dias &nbsp;|&nbsp; v5.0</div>

<div class="criticidad-wrap">
  <div class="criticidad-label">🔥 Nivel de criticidad global ({cnt['VENCIDO']} vencidos + {cnt['PROXIMO']} proximos sobre {total_certs} certificados evaluados)</div>
  <div class="criticidad-bar-bg">
    <div class="criticidad-bar" id="crit-bar" style="width:{pct_critico}%"></div>
  </div>
  <div class="criticidad-pct">{pct_critico}% en estado critico o de alerta</div>
</div>

<div class="cards">
  <div class="card rojo"  onclick="filtrarEstado('VENCIDO')">
    <div class="num">{cnt['VENCIDO']}</div><div class="lbl">🔴 Vencidos</div>
  </div>
  <div class="card amari" onclick="filtrarEstado('PROXIMO')">
    <div class="num">{cnt['PROXIMO']}</div><div class="lbl">🟡 Proximos a vencer</div>
  </div>
  <div class="card azul"  onclick="filtrarEstado('ACTUALIZADO')">
    <div class="num">{cnt['ACTUALIZADO']}</div><div class="lbl">🔵 Datos actualizados</div>
  </div>
  <div class="card gris"  onclick="filtrarEstado('SIN_ARCHIVO')">
    <div class="num">{cnt['SIN_ARCHIVO']}</div><div class="lbl">⚪ Sin archivo .out</div>
  </div>
</div>

<div class="filtros">
  <label>Filtrar:</label>
  <select id="fil-estado" onchange="cambioFiltro()">
    <option value="">Todos los estados</option>
    <option value="VENCIDO">🔴 Vencidos</option>
    <option value="PROXIMO">🟡 Proximos a vencer</option>
    <option value="ACTUALIZADO">🔵 Actualizados</option>
    <option value="SIN_ARCHIVO">⚪ Sin archivo</option>
  </select>
  <select id="fil-ambiente" onchange="cambioFiltro()">
    <option value="">Todos los ambientes</option>
    {''.join(f'<option value="{a}">{a}</option>' for a in AMBIENTES)}
  </select>
  <select id="fil-tipo" onchange="cambioFiltro()">
    <option value="">Todos los tipos</option>
    <option value="WAS">WAS</option>
    <option value="AIPAC">AIPAC</option>
    <option value="PLUG">PLUG.WAS</option>
  </select>
  <input type="text" id="fil-buscar" placeholder="🔍 Buscar alias..." oninput="cambioFiltro()">
  <button class="btn btn-reset" onclick="resetFiltros()">✕ Limpiar</button>
  <button class="btn btn-csv"   onclick="exportarCSV()">⬇ Exportar CSV</button>
</div>

<div class="tabla-wrap">
  <table>
    <thead>
      <tr>
        <th onclick="ordenar('ambiente')">Ambiente ↕</th>
        <th onclick="ordenar('hoja')">Hoja ↕</th>
        <th onclick="ordenar('alias')">Alias ↕</th>
        <th onclick="ordenar('estado')">Estado ↕</th>
        <th onclick="ordenar('dias')">Dias restantes ↕</th>
        <th onclick="ordenar('fecha')">Vencimiento ↕</th>
        <th>Detalle</th>
      </tr>
    </thead>
    <tbody id="tbody"></tbody>
  </table>
  <div class="no-rows" id="no-rows" style="display:none">Sin resultados.</div>
</div>

<div class="paginacion" id="paginacion"></div>
<div class="footer">Mostrando <span id="cnt-visible">0</span> de {len(registros)} registros</div>

<script>
const datos = {datos_json};
const POR_PAGINA = 25;
let orden = {{col: 'dias', asc: true}};
let paginaActual = 1;
let filasFiltradas = [];

function filtrarEstado(e) {{
  document.getElementById('fil-estado').value = e;
  cambioFiltro();
}}

function resetFiltros() {{
  document.getElementById('fil-estado').value   = '';
  document.getElementById('fil-ambiente').value = '';
  document.getElementById('fil-tipo').value     = '';
  document.getElementById('fil-buscar').value   = '';
  cambioFiltro();
}}

function cambioFiltro() {{
  paginaActual = 1;
  renderizar();
}}

function badge(e) {{
  const labels = {{VENCIDO:'🔴 VENCIDO', PROXIMO:'🟡 POR VENCER',
                   ACTUALIZADO:'🔵 ACTUALIZADO', SIN_ARCHIVO:'⚪ SIN ARCHIVO'}};
  return '<span class="badge badge-' + e + '">' + (labels[e]||e) + '</span>';
}}

function diasHtml(d, estado) {{
  if (estado === 'ACTUALIZADO' || estado === 'SIN_ARCHIVO') return '<span style="color:#bbb">—</span>';
  if (d < 0) return '<span class="dias-critico">Vencido hace ' + Math.abs(d) + ' dias</span>';
  if (d <= {dias_alerta}) return '<span class="dias-alerta">' + d + ' dias</span>';
  return '<span class="dias-ok">' + d + ' dias</span>';
}}

function aplicarFiltros() {{
  const fe = document.getElementById('fil-estado').value;
  const fa = document.getElementById('fil-ambiente').value;
  const ft = document.getElementById('fil-tipo').value;
  const fb = document.getElementById('fil-buscar').value.toLowerCase();
  return datos.filter(d => {{
    if (fe && d.estado !== fe) return false;
    if (fa && d.ambiente !== fa) return false;
    if (ft && !d.hoja.includes(ft)) return false;
    if (fb && !d.alias.toLowerCase().includes(fb) && !d.hoja.toLowerCase().includes(fb)) return false;
    return true;
  }});
}}

function ordenar(col) {{
  orden = {{col, asc: orden.col === col ? !orden.asc : true}};
  renderizar();
}}

function renderizarPaginacion(total) {{
  const totalPags = Math.ceil(total / POR_PAGINA);
  const wrap = document.getElementById('paginacion');
  if (totalPags <= 1) {{ wrap.innerHTML = ''; return; }}

  let html = '';
  html += '<button onclick="irPagina(' + Math.max(1, paginaActual-1) + ')" ' + (paginaActual===1?'disabled':'') + '>‹ Anterior</button>';

  // Paginas con elipsis
  let inicio = Math.max(1, paginaActual - 2);
  let fin    = Math.min(totalPags, paginaActual + 2);
  if (inicio > 1) html += '<button onclick="irPagina(1)">1</button>' + (inicio > 2 ? '<span class="info">…</span>' : '');
  for (let i = inicio; i <= fin; i++) {{
    html += '<button class="' + (i===paginaActual?'activa':'') + '" onclick="irPagina(' + i + ')">' + i + '</button>';
  }}
  if (fin < totalPags) html += (fin < totalPags-1 ? '<span class="info">…</span>' : '') + '<button onclick="irPagina(' + totalPags + ')">' + totalPags + '</button>';

  html += '<button onclick="irPagina(' + Math.min(totalPags, paginaActual+1) + ')" ' + (paginaActual===totalPags?'disabled':'') + '>Siguiente ›</button>';
  html += '<span class="info">Pagina ' + paginaActual + ' de ' + totalPags + '</span>';
  wrap.innerHTML = html;
}}

function irPagina(p) {{
  paginaActual = p;
  renderizar();
}}

function renderizar() {{
  filasFiltradas = aplicarFiltros();
  filasFiltradas.sort((a, b) => {{
    let va = a[orden.col], vb = b[orden.col];
    if (orden.col === 'dias') {{ va = Number(va); vb = Number(vb); return orden.asc ? va-vb : vb-va; }}
    return orden.asc ? String(va).localeCompare(String(vb)) : String(vb).localeCompare(String(va));
  }});

  const inicio  = (paginaActual - 1) * POR_PAGINA;
  const pagina  = filasFiltradas.slice(inicio, inicio + POR_PAGINA);

  const tbody = document.getElementById('tbody');
  document.getElementById('no-rows').style.display = filasFiltradas.length ? 'none' : 'block';
  document.getElementById('cnt-visible').textContent = filasFiltradas.length;

  tbody.innerHTML = pagina.map(d => `
    <tr class="row-${{d.estado}}">
      <td><strong>${{d.ambiente}}</strong></td>
      <td>${{d.hoja || '—'}}</td>
      <td style="max-width:280px;word-break:break-word">${{d.alias}}</td>
      <td>${{badge(d.estado)}}</td>
      <td>${{diasHtml(d.dias, d.estado)}}</td>
      <td>${{d.fecha === '-' ? '<span style="color:#bbb">—</span>' : d.fecha}}</td>
      <td style="font-size:12px;color:#555">${{d.detalle}}</td>
    </tr>`).join('');

  renderizarPaginacion(filasFiltradas.length);
}}

function exportarCSV() {{
  const filas = filasFiltradas.length ? filasFiltradas : aplicarFiltros();
  const encabezado = ['Ambiente','Hoja','Alias','Estado','Dias restantes','Vencimiento','Detalle'];
  const lineas = [encabezado.join(';')];
  filas.forEach(d => {{
    const diasVal = (d.estado === 'ACTUALIZADO' || d.estado === 'SIN_ARCHIVO') ? '-' : d.dias;
    lineas.push([d.ambiente, d.hoja, '"'+d.alias.replace(/"/g,'""')+'"', d.estado, diasVal, d.fecha, d.detalle].join(';'));
  }});
  const blob = new Blob(['\ufeff' + lineas.join('\\n')], {{type:'text/csv;charset=utf-8;'}});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href     = url;
  a.download = 'auditoria_ssl_{fecha_ejecucion}.csv';
  a.click();
  URL.revokeObjectURL(url);
}}

renderizar();
</script>
</body>
</html>"""

    with open(archivo_html, "w", encoding="utf-8") as f:
        f.write(html)
    log("Reporte HTML generado: " + os.path.basename(archivo_html))


# ==========================
# PROCESO PRINCIPAL
# ==========================
_diffs_set = set()  # Para deduplicar alertas en diffs


def ejecutar_proceso():
    os.makedirs(RAIZ, exist_ok=True)

    EXCEL_OUT = _nombre_excel_salida()

    if os.path.exists(ARCHIVO_LOG):
        os.remove(ARCHIVO_LOG)
    if os.path.exists(LOG_VENCIMIENTOS):
        os.remove(LOG_VENCIMIENTOS)

    log("=" * 60)
    log("  AUDITORIA SSL v5.0 - INICIO (" + str(date.today()) + ")")
    log("  Alerta amarilla: certificados que vencen en " + str(DIAS_ALERTA) + " dias o menos")
    log("  Archivo de salida: " + os.path.basename(EXCEL_OUT))
    log("=" * 60)

    for amb in AMBIENTES:
        c = carpeta_ambiente(amb)
        n = len(os.listdir(c)) if os.path.exists(c) else 0
        log("  " + amb + ": " + str(n) + " archivos en " + c)

    if not os.path.exists(EXCEL_IN):
        log("ERROR: No se encontro: " + EXCEL_IN, "ERROR")
        return

    wb    = load_workbook(EXCEL_IN)
    diffs = []

    for sheet_name in wb.sheetnames:
        ws      = wb[sheet_name]
        nombre  = sheet_name.upper()
        ambiente = next((a for a in AMBIENTES if a in nombre), None)

        if not ambiente:
            log("Hoja '" + sheet_name + "': sin ambiente, se omite.", "WARN")
            continue

        log("Hoja: " + sheet_name + " | Ambiente: " + ambiente)

        if "PLUG.WAS" in nombre:
            procesar_hoja_plug_was(ws, ambiente, diffs)
        elif nombre.endswith("WAS"):
            procesar_hoja_was(ws, ambiente, diffs)
        elif nombre.endswith("AIPAC"):
            procesar_hoja_aipac(ws, ambiente, diffs)
        else:
            log("  '" + sheet_name + "': tipo no reconocido.", "WARN")

    log("Guardando en: " + EXCEL_OUT)
    wb.save(EXCEL_OUT)

    alertas = [d for d in diffs if "VENCIDO" in d or "VENCER" in d]
    cambios = [d for d in diffs if "VENCIDO" not in d and "VENCER" not in d]

    log("\n" + "=" * 60)
    log("  RESUMEN - CAMBIOS Y ALERTAS DE VENCIMIENTO")
    log("=" * 60)
    if alertas:
        log("  ALERTAS DE VENCIMIENTO (" + str(len(alertas)) + "):")
        for a in alertas:
            log("    " + a, "ALERT")
    if cambios:
        log("  DATOS ACTUALIZADOS (" + str(len(cambios)) + "):")
        for c in cambios:
            log("    " + c, "CAMBIO")
    if not diffs:
        log("  Sin diferencias ni alertas. Todo OK.")
    log("\n>>> PROCESO FINALIZADO <<<")

    # Estadisticas de cobertura
    imprimir_estadisticas()

    # Log separado de vencimientos
    ts       = date.today().strftime("%d/%m/%Y")
    vencidos = [a for a in alertas if "VENCIDO" in a]
    proximos = [a for a in alertas if "VENCER" in a]

    with open(LOG_VENCIMIENTOS, "w", encoding="utf-8") as f:
        f.write("=" * 60 + "\n")
        f.write("  REPORTE DE VENCIMIENTO DE CERTIFICADOS\n")
        f.write("  Generado: " + ts + "\n")
        f.write("  Umbral de alerta: " + str(DIAS_ALERTA) + " dias\n")
        f.write("=" * 60 + "\n\n")

        if vencidos:
            f.write("CERTIFICADOS VENCIDOS (" + str(len(vencidos)) + "):\n")
            f.write("-" * 40 + "\n")
            for v in vencidos:
                f.write("  " + v + "\n")
            f.write("\n")

        if proximos:
            f.write("PROXIMOS A VENCER - menos de " + str(DIAS_ALERTA) + " dias (" + str(len(proximos)) + "):\n")
            f.write("-" * 40 + "\n")
            for p in proximos:
                f.write("  " + p + "\n")
            f.write("\n")

        if not alertas:
            f.write("  Sin alertas. Todos los certificados estan vigentes.\n")

    log("Log de vencimientos guardado en: " + LOG_VENCIMIENTOS)

    # Generar reporte HTML
    generar_html_reporte(ARCHIVO_LOG, HTML_REPORTE, str(date.today()), DIAS_ALERTA)


if __name__ == "__main__":
    ejecutar_proceso()